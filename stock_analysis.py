import openpyxl
from openpyxl.utils import get_column_letter
import yfinance as yf
from datetime import datetime

# Function to fetch stock data from Yahoo Finance API
def fetch_stock_data(stock_symbol, start_date, end_date):
    stock = yf.Ticker(stock_symbol)
    data = stock.history(start=start_date, end=end_date)
    if data.empty:
        raise ValueError(f"No data found for {stock_symbol} between {start_date} and {end_date}")
    data = data[::-1]  # Reverse the date order to get chronological sequence
    return data[['High', 'Low']].reset_index()

# Main analysis function that creates Excel file with stock analysis
def create_stock_analysis_file(stock_symbol, start_date, end_date, output_file):
    # Fetch stock data
    data = fetch_stock_data(stock_symbol, start_date, end_date)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{stock_symbol} Analysis"

    # Set up headers
    headers = ["Date", "High", "Low", "Change"]
    for col, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col)}1"] = header

    # Fill in the data
    for row, (date, high, low) in enumerate(zip(data['Date'], data['High'], data['Low']), 2):
        high = round(float(high), 2)  # Round to 2 decimal places
        low = round(float(low), 2)    # Round to 2 decimal places
        change = round(high - low, 2) # Calculate and round price change
        ws[f"A{row}"] = date.strftime('%Y-%m-%d')
        ws[f"B{row}"] = high
        ws[f"C{row}"] = low
        ws[f"D{row}"] = change

    # Define quadrant values for supply-demand analysis
    positive_values = [0.25 * i for i in range(1, 17)]  # 0.25, 0.50, ..., 4.0
    negative_values = [-0.25 * i for i in range(1, 17)] # -0.25, -0.50, ..., -4.0

    # Set up Positive Quadrant headers (values only, no title)
    for col, value in enumerate(positive_values, 5):  # Start from column E (5)
        ws[f"{get_column_letter(col)}1"] = value

    # Set up Negative Quadrant headers (values only, no title)
    for col, value in enumerate(negative_values, 21): # Start from column U (21)
        ws[f"{get_column_letter(col)}1"] = value

    # Fill in Positive Quadrant data: High + (Change * Quadrant)
    for row in range(2, len(data) + 2):  # Start from row 2 (row 1 is header)
        high = float(ws[f"B{row}"].value)
        change = float(ws[f"D{row}"].value)
        for col, quadrant in enumerate(positive_values, 5):  # Start from column E
            value = high + (change * quadrant)
            ws[f"{get_column_letter(col)}{row}"] = round(value, 2)

    # Fill in Negative Quadrant data: Low + (Change * Quadrant)
    for row in range(2, len(data) + 2):  # Start from row 2
        low = float(ws[f"C{row}"].value)
        change = float(ws[f"D{row}"].value)
        for col, quadrant in enumerate(negative_values, 21):  # Start from column U
            value = low + (change * quadrant)
            ws[f"{get_column_letter(col)}{row}"] = round(value, 2)

    # Save the file
    wb.save(output_file)
    print(f"File saved as {output_file}")

# Main execution block
if __name__ == "__main__":
    # Get user input for analysis parameters
    print("You can find valid stock symbols from Yahoo Finance (e.g., AAPL for Apple, ^GSPC for S&P 500, EURUSD=X for EUR/USD)")
    stock_symbol = input("Enter stock symbol (e.g., AAPL for Apple, ^GSPC for S&P 500): ").upper()
    start_date = input("Enter start date (YYYY-MM-DD): ")
    end_date = min(datetime.strptime(input("Enter end date (YYYY-MM-DD): "), "%Y-%m-%d"), datetime(2025, 3, 15))
    
    # Get output file name and ensure it has .xlsx extension
    output_file = input("Enter the output file name (e.g., stock_analysis.xlsx): ")
    if not output_file.endswith('.xlsx'):
        output_file += '.xlsx'
    
    try:
        # Validate dates and run analysis
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        if start_date >= end_date:
            raise ValueError("Start date must be before end date.")
        create_stock_analysis_file(stock_symbol, start_date, end_date, output_file)
    except ValueError as e:
        print(f"Error: {e}. Please enter dates in YYYY-MM-DD format.")